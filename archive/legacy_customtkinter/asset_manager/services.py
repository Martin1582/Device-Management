import csv
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from .constants import DEVICE_TYPES, STATUS_ACTIVE, STATUS_INACTIVE


HEADER_ALIASES = {
    "type": {
        "typ",
        "type",
        "geraetetyp",
        "geratetyp",
        "geraettyp",
        "gerattyp",
        "device type",
        "device_type",
        "kategorie",
    },
    "user_name": {
        "user",
        "user name",
        "username",
        "benutzer",
        "mitarbeiter",
        "mitarbeitername",
        "employee",
        "name",
    },
    "model": {
        "modell",
        "model",
        "geraete modell",
        "geraetemodell",
        "gerate modell",
        "device model",
        "device",
        "geraet",
        "gerat",
    },
    "asset_tag": {
        "s/n / imei",
        "s/n",
        "sn",
        "serial",
        "serial number",
        "seriennummer",
        "imei",
        "asset tag",
        "assettag",
        "asset-tag",
    },
    "extra_info": {
        "rufnummer / hostname",
        "hostname",
        "host",
        "extra info",
        "extra_info",
    },
    "status": {"status"},
}


def normalize_header(value):
    text = str(value or "").strip().lower()
    replacements = {
        "ä": "ae",
        "ö": "oe",
        "ü": "ue",
        "ß": "ss",
        "_": " ",
        "-": " ",
        "\n": " ",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return " ".join(text.split())


def normalize_asset_tag(value):
    return str(value or "").strip().upper()


def normalize_extra_info(value):
    return str(value or "").strip()


def normalize_device_type(value, default_type=None):
    text = normalize_header(value)
    if not text and default_type:
        return default_type
    if text in {"smartphone", "phone", "handy", "mobiltelefon"}:
        return "Smartphone"
    if text in {"notebook", "laptop"}:
        return "Notebook"
    if value in DEVICE_TYPES:
        return value
    raise ValueError(f"Unbekannter Gerätetyp: {value}")


def normalize_status(value):
    text = normalize_header(value)
    if not text:
        return STATUS_ACTIVE
    if text in {"aktiv", "active"}:
        return STATUS_ACTIVE
    if text in {"inaktiv", "inactive"}:
        return STATUS_INACTIVE
    raise ValueError(f"Unbekannter Status: {value}")


def map_headers(headers):
    mapped = {}
    for index, header in enumerate(headers):
        normalized = normalize_header(header)
        for canonical_name, aliases in HEADER_ALIASES.items():
            if normalized in aliases and canonical_name not in mapped:
                mapped[canonical_name] = index
                break
    return mapped


def extract_cell(row, header_map, field_name):
    if field_name not in header_map:
        return ""
    value = row[header_map[field_name]]
    return "" if value is None else str(value).strip()


def parse_asset_rows_from_workbook(file_path, default_type=None):
    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        raise ValueError("Die Excel-Datei enthält keine Daten.")

    header_map = map_headers(rows[0])
    required_headers = {"user_name", "model", "asset_tag"}
    missing_headers = sorted(required_headers - set(header_map))
    if missing_headers:
        raise ValueError(
            "Folgende Spalten fehlen: "
            + ", ".join(missing_headers)
            + ". Erwartet werden mindestens User, Modell und S/N / IMEI."
        )

    parsed_rows = []
    row_errors = []
    seen_asset_tags = set()
    seen_extra_info = set()

    for row_number, row in enumerate(rows[1:], start=2):
        if not any(value not in (None, "") for value in row):
            continue

        try:
            device_type = normalize_device_type(extract_cell(row, header_map, "type"), default_type=default_type)
            user_name = extract_cell(row, header_map, "user_name")
            model = extract_cell(row, header_map, "model")
            asset_tag = normalize_asset_tag(extract_cell(row, header_map, "asset_tag"))
            extra_info = normalize_extra_info(extract_cell(row, header_map, "extra_info"))
            status = normalize_status(extract_cell(row, header_map, "status"))
            if device_type == "Smartphone":
                extra_info = ""

            missing_values = [
                label
                for label, value in (
                    ("User", user_name),
                    ("Modell", model),
                    ("S/N / IMEI", asset_tag),
                )
                if not value
            ]
            if missing_values:
                raise ValueError(f"Pflichtfelder fehlen: {', '.join(missing_values)}")
            if asset_tag in seen_asset_tags:
                raise ValueError(f"Doppeltes Asset-Tag in der Excel-Datei: {asset_tag}")
            seen_asset_tags.add(asset_tag)
            if extra_info:
                extra_key = (device_type, extra_info.casefold())
                if extra_key in seen_extra_info:
                    raise ValueError(f"Doppelte Zusatzinformation in der Excel-Datei: {extra_info}")
                seen_extra_info.add(extra_key)

            parsed_rows.append(
                {
                    "type": device_type,
                    "user_name": user_name,
                    "model": model,
                    "asset_tag": asset_tag,
                    "extra_info": extra_info,
                    "status": status,
                }
            )
        except ValueError as exc:
            row_errors.append(f"Zeile {row_number}: {exc}")

    if not parsed_rows and row_errors:
        raise ValueError("\n".join(row_errors[:10]))

    return parsed_rows, row_errors


def build_import_preview(file_path, default_type=None):
    parsed_rows, row_errors = parse_asset_rows_from_workbook(file_path, default_type=default_type)
    preview_rows = parsed_rows[:10]
    counts = {
        "total_valid": len(parsed_rows),
        "total_errors": len(row_errors),
        "by_type": {device_type: 0 for device_type in DEVICE_TYPES},
    }
    for row in parsed_rows:
        counts["by_type"][row["type"]] += 1
    return {
        "rows": preview_rows,
        "counts": counts,
        "errors": row_errors[:10],
    }


def import_assets_from_workbook(file_path, db_manager, default_type=None, actor="Import"):
    parsed_rows, row_errors = parse_asset_rows_from_workbook(file_path, default_type=default_type)
    summary = {"created": 0, "updated": 0, "skipped": len(row_errors), "errors": row_errors}

    for row in parsed_rows:
        result = db_manager.upsert_asset(
            row["type"],
            row["user_name"],
            row["model"],
            row["asset_tag"],
            row["extra_info"],
            row["status"],
            actor=actor,
        )
        summary[result] += 1

    return summary


def export_assets_to_csv(rows, file_path):
    with open(file_path, mode="w", newline="", encoding="utf-8-sig") as file_handle:
        writer = csv.writer(file_handle, delimiter=";")
        writer.writerow(["Typ", "User", "Modell", "S/N / IMEI", "Hostname", "Status", "Zuletzt geändert", "Geändert von"])
        for row in rows:
            writer.writerow(
                [
                    row["type"],
                    row["user_name"],
                    row["model"],
                    row["asset_tag"],
                    row["extra_info"],
                    row["status"],
                    row["updated_at"],
                    row["updated_by"],
                ]
            )


def export_assets_to_printable_html(rows, file_path, title="Inventarliste"):
    html_rows = []
    for row in rows:
        html_rows.append(
            "<tr>"
            f"<td>{row['type']}</td>"
            f"<td>{row['user_name']}</td>"
            f"<td>{row['model']}</td>"
            f"<td>{row['asset_tag']}</td>"
            f"<td>{row['extra_info']}</td>"
            f"<td>{row['status']}</td>"
            "</tr>"
        )
    html = f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>{title}</title>
  <style>
    body {{ font-family: Arial, sans-serif; margin: 32px; }}
    table {{ width: 100%; border-collapse: collapse; }}
    th, td {{ border: 1px solid #c9d2dc; padding: 8px; text-align: left; }}
    th {{ background: #eef3f8; }}
  </style>
</head>
<body>
  <h1>{title}</h1>
  <p>Stand: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}</p>
  <table>
    <thead>
      <tr><th>Typ</th><th>User</th><th>Modell</th><th>S/N / IMEI</th><th>Hostname</th><th>Status</th></tr>
    </thead>
    <tbody>
      {''.join(html_rows)}
    </tbody>
  </table>
</body>
</html>
"""
    Path(file_path).write_text(html, encoding="utf-8")


def build_duplicate_report(rows):
    issues = []
    seen_asset_tags = set()
    seen_hostnames = set()
    for row in rows:
        asset_tag = normalize_asset_tag(row["asset_tag"])
        if asset_tag in seen_asset_tags:
            issues.append(f"Doppeltes Asset-Tag: {asset_tag}")
        seen_asset_tags.add(asset_tag)

        if row["type"] == "Notebook" and row["extra_info"]:
            host_key = row["extra_info"].strip().casefold()
            if host_key in seen_hostnames:
                issues.append(f"Doppelter Hostname: {row['extra_info']}")
            seen_hostnames.add(host_key)
        if row["type"] == "Notebook" and not row["extra_info"]:
            issues.append(f"Notebook ohne Hostname: {row['asset_tag']}")
    return issues


def build_export_filename():
    date_str = datetime.now().strftime("%Y-%m-%d")
    return f"Brunel_Assets_{date_str}.csv"


def build_backup_filename():
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    return f"DeviceManagement_Backup_{date_str}.db"


def build_print_filename():
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    return f"Inventarliste_{date_str}.html"


def build_import_filename(file_path):
    return Path(file_path).name


def write_import_protocol(file_path, summary, default_type=None):
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    protocol_path = Path(file_path).with_name(f"Importprotokoll_{timestamp}.txt")
    lines = [
        "Device Management - Importprotokoll",
        f"Quelle: {build_import_filename(file_path)}",
        f"Zeitpunkt: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}",
        f"Neu angelegt: {summary['created']}",
        f"Aktualisiert: {summary['updated']}",
        f"Übersprungen: {summary['skipped']}",
    ]
    if default_type:
        lines.extend(
            [
                "",
                "Hinweis:",
                f"Falls keine Typ-Spalte vorhanden war, wurde '{default_type}' als Standard verwendet.",
            ]
        )
    if summary["errors"]:
        lines.extend(["", "Details:"])
        lines.extend(summary["errors"])
    protocol_path.write_text("\n".join(lines), encoding="utf-8")
    return protocol_path


def create_backup(source_path, destination_path):
    source = Path(source_path)
    destination = Path(destination_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, destination)
    return destination


def restore_backup(source_path, destination_path):
    source = Path(source_path)
    destination = Path(destination_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, destination)
    return destination
