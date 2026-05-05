from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from ..db.migrations import DEVICE_TYPE_VALUES, INVENTORY_STATUS_VALUES, normalize_asset_tag
from ..db.repository import DatabaseRepository


HEADER_ALIASES = {
    "device_type": {
        "typ",
        "type",
        "geraetetyp",
        "geratetyp",
        "geraettyp",
        "gerattyp",
        "device type",
        "device_type",
        "kategorie",
    },
    "asset_tag": {
        "s/n / imei",
        "s/n",
        "sn",
        "serial",
        "serial number",
        "seriennummer",
        "imei",
        "asset tag",
        "assettag",
        "asset-tag",
        "sn / imei",
    },
    "model_name": {
        "modell",
        "model",
        "geraete modell",
        "geraetemodell",
        "gerate modell",
        "device model",
        "device",
        "geraet",
        "gerat",
    },
    "manufacturer": {"hersteller", "manufacturer", "vendor", "marke"},
    "inventory_status": {"status", "inventarstatus", "inventory status", "inventory_status"},
    "assigned_to": {
        "user",
        "user name",
        "username",
        "benutzer",
        "mitarbeiter",
        "mitarbeitername",
        "employee",
        "name",
        "person",
    },
    "hostname": {"rufnummer / hostname", "hostname", "host", "extra info", "extra_info"},
    "source_asset_tag": {"quelle", "alt-id", "alt id", "legacy id", "source", "source_asset_tag"},
    "notes": {"notizen", "notes", "bemerkung", "bemerkungen", "kommentar"},
}


@dataclass(frozen=True)
class ImportAssetRow:
    row_number: int
    device_type: str = ""
    asset_tag: str = ""
    model_name: str = ""
    manufacturer: str = ""
    inventory_status: str = "active"
    assigned_to: str = ""
    hostname: str = ""
    source_asset_tag: str = ""
    notes: str = ""
    status: str = "error"
    messages: tuple[str, ...] = field(default_factory=tuple)

    @property
    def can_import(self) -> bool:
        return self.status == "ok"

    @property
    def message_text(self) -> str:
        return "; ".join(self.messages)


@dataclass(frozen=True)
class ImportPreview:
    source_path: Path
    rows: tuple[ImportAssetRow, ...]

    @property
    def total_count(self) -> int:
        return len(self.rows)

    @property
    def importable_count(self) -> int:
        return sum(1 for row in self.rows if row.can_import)

    @property
    def error_count(self) -> int:
        return sum(1 for row in self.rows if row.status == "error")

    @property
    def duplicate_count(self) -> int:
        return sum(1 for row in self.rows if row.status == "duplicate")

    @property
    def skipped_count(self) -> int:
        return self.total_count - self.importable_count


@dataclass(frozen=True)
class ImportSummary:
    created_assets: int
    created_assignments: int
    skipped: int
    errors: tuple[str, ...] = field(default_factory=tuple)


def build_import_preview(file_path: str | Path, repository: DatabaseRepository) -> ImportPreview:
    path = Path(file_path)
    raw_rows = _read_tabular_file(path)
    if not raw_rows:
        raise ValueError("Die Importdatei enthaelt keine Daten.")

    header_map = _map_headers(raw_rows[0])
    required_headers = {"asset_tag", "model_name"}
    missing_headers = sorted(required_headers - set(header_map))
    if missing_headers:
        raise ValueError(
            "Folgende Pflichtspalten fehlen: "
            + ", ".join(missing_headers)
            + ". Erwartet werden mindestens Modell und SN / IMEI."
        )

    existing_asset_tags = {
        normalize_asset_tag(row["asset_tag"])
        for row in repository.list_assets()
        if row.get("asset_tag")
    }
    active_hostnames = {
        str(row["hostname"]).strip().casefold()
        for row in repository.list_asset_snapshots()
        if row.get("hostname")
    }

    seen_asset_tags: set[str] = set()
    seen_hostnames: set[str] = set()
    preview_rows = []
    for row_number, raw_row in enumerate(raw_rows[1:], start=2):
        if _is_empty_row(raw_row):
            continue
        parsed = _parse_import_row(raw_row, row_number, header_map)
        messages = list(parsed.messages)
        asset_key = normalize_asset_tag(parsed.asset_tag)
        hostname_key = parsed.hostname.casefold()

        if not parsed.device_type:
            messages.append("Typ fehlt oder ist ungueltig.")
        if not parsed.asset_tag:
            messages.append("SN / IMEI fehlt.")
        if not parsed.model_name:
            messages.append("Modell fehlt.")
        if parsed.inventory_status not in INVENTORY_STATUS_VALUES:
            messages.append("Inventarstatus ist ungueltig.")
        if asset_key and asset_key in seen_asset_tags:
            messages.append("SN / IMEI ist in der Importdatei doppelt.")
        if asset_key and asset_key in existing_asset_tags:
            messages.append("SN / IMEI existiert bereits in der Datenbank.")
        if parsed.device_type == "Notebook" and hostname_key:
            if hostname_key in seen_hostnames:
                messages.append("Hostname ist in der Importdatei doppelt.")
            if hostname_key in active_hostnames:
                messages.append("Hostname ist bereits aktiv zugewiesen.")

        if asset_key:
            seen_asset_tags.add(asset_key)
        if parsed.device_type == "Notebook" and hostname_key:
            seen_hostnames.add(hostname_key)

        status = _status_from_messages(messages)
        preview_rows.append(_replace_status(parsed, status=status, messages=tuple(messages)))

    return ImportPreview(source_path=path, rows=tuple(preview_rows))


def import_preview_rows(
    preview: ImportPreview,
    repository: DatabaseRepository,
    *,
    actor: str = "pyside-import",
) -> ImportSummary:
    created_assets = 0
    created_assignments = 0
    errors = []

    for row in preview.rows:
        if not row.can_import:
            continue
        try:
            asset = repository.create_asset(
                row.device_type,
                row.asset_tag,
                row.model_name,
                manufacturer=row.manufacturer,
                inventory_status=row.inventory_status,
                notes=row.notes,
                source_asset_tag=row.source_asset_tag,
            )
            created_assets += 1
            if row.assigned_to or row.hostname:
                person_id = None
                if row.assigned_to:
                    person = repository.create_or_update_person(row.assigned_to)
                    person_id = person["id"]
                repository.assign_asset(
                    asset_id=asset["id"],
                    person_id=person_id,
                    hostname=row.hostname,
                    notes="Automatisch aus Importdatei uebernommen.",
                    actor=actor,
                )
                created_assignments += 1
        except Exception as exc:
            errors.append(f"Zeile {row.row_number}: {exc}")

    return ImportSummary(
        created_assets=created_assets,
        created_assignments=created_assignments,
        skipped=preview.skipped_count + len(errors),
        errors=tuple(errors),
    )


def _read_tabular_file(path: Path) -> list[list[str]]:
    if not path.exists():
        raise ValueError(f"Importdatei nicht gefunden: {path}")
    suffix = path.suffix.casefold()
    if suffix == ".xlsx":
        return _read_xlsx(path)
    if suffix == ".csv":
        return _read_csv(path)
    raise ValueError("Unterstuetzt werden aktuell CSV- und XLSX-Dateien.")


def _read_xlsx(path: Path) -> list[list[str]]:
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        return [[_cell(value) for value in row] for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _read_csv(path: Path) -> list[list[str]]:
    sample = path.read_text(encoding="utf-8-sig")[:2048]
    delimiter = ";"
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,")
        delimiter = dialect.delimiter
    except csv.Error:
        pass
    with path.open("r", encoding="utf-8-sig", newline="") as file_handle:
        return [[_cell(value) for value in row] for row in csv.reader(file_handle, delimiter=delimiter)]


def _parse_import_row(row: list[str], row_number: int, header_map: dict[str, int]) -> ImportAssetRow:
    device_type = _normalize_device_type(_extract(row, header_map, "device_type"))
    inventory_status = _normalize_inventory_status(_extract(row, header_map, "inventory_status"))
    hostname = _extract(row, header_map, "hostname")
    if device_type == "Smartphone":
        hostname = ""
    return ImportAssetRow(
        row_number=row_number,
        device_type=device_type,
        asset_tag=normalize_asset_tag(_extract(row, header_map, "asset_tag")),
        model_name=_extract(row, header_map, "model_name"),
        manufacturer=_extract(row, header_map, "manufacturer"),
        inventory_status=inventory_status,
        assigned_to=_extract(row, header_map, "assigned_to"),
        hostname=hostname,
        source_asset_tag=_extract(row, header_map, "source_asset_tag"),
        notes=_extract(row, header_map, "notes"),
    )


def _map_headers(headers: Iterable[str]) -> dict[str, int]:
    mapped = {}
    for index, header in enumerate(headers):
        normalized = _normalize_header(header)
        for field_name, aliases in HEADER_ALIASES.items():
            if normalized in aliases and field_name not in mapped:
                mapped[field_name] = index
                break
    return mapped


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().casefold()
    replacements = {
        "ä": "ae",
        "ö": "oe",
        "ü": "ue",
        "ß": "ss",
        "_": " ",
        "-": " ",
        "\n": " ",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return " ".join(text.split())


def _normalize_device_type(value: str) -> str:
    text = _normalize_header(value)
    if text in {"smartphone", "phone", "handy", "mobiltelefon"}:
        return "Smartphone"
    if text in {"notebook", "laptop"}:
        return "Notebook"
    if value in DEVICE_TYPE_VALUES:
        return value
    return ""


def _normalize_inventory_status(value: str) -> str:
    text = _normalize_header(value)
    if not text or text in {"aktiv", "active"}:
        return "active"
    if text in {"inaktiv", "inactive"}:
        return "inactive"
    if text in {"retired", "ausgemustert", "stillgelegt"}:
        return "retired"
    return value.strip()


def _extract(row: list[str], header_map: dict[str, int], field_name: str) -> str:
    index = header_map.get(field_name)
    if index is None or index >= len(row):
        return ""
    return _cell(row[index])


def _cell(value: object) -> str:
    return "" if value is None else str(value).strip()


def _is_empty_row(row: list[str]) -> bool:
    return not any(_cell(value) for value in row)


def _status_from_messages(messages: list[str]) -> str:
    if not messages:
        return "ok"
    if any("existiert bereits" in message for message in messages):
        return "duplicate"
    return "error"


def _replace_status(row: ImportAssetRow, *, status: str, messages: tuple[str, ...]) -> ImportAssetRow:
    return ImportAssetRow(
        row_number=row.row_number,
        device_type=row.device_type,
        asset_tag=row.asset_tag,
        model_name=row.model_name,
        manufacturer=row.manufacturer,
        inventory_status=row.inventory_status,
        assigned_to=row.assigned_to,
        hostname=row.hostname,
        source_asset_tag=row.source_asset_tag,
        notes=row.notes,
        status=status,
        messages=messages,
    )
